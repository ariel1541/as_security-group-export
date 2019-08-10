import boto3
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side

# Variables
awsAccID = '123456789012' # AWS account number
assumeRole = 'arn:' # ARN of role that you want to assume on account above
vpcId = 'vpc-' # VPC ID from account above
fileName = 'SecurityGroups - AWS ' + awsAccID + '.xlsx' # Name of .xlsx file
tokenSerial = '' # In case of MFA - Serial number of hardware token

# Assume role on specific account and get back ec2 client credentials
def login(awsAccID, prefix):
    token = input('Enter token code: ')
    clientSTS = boto3.client('sts')
    stsCreds = clientSTS.assume_role(
                RoleArn=assumeRole,
                RoleSessionName='SecurityGroupsExport',
                DurationSeconds=3600,
                SerialNumber=tokenSerial,
                TokenCode=str(token)
            )

    client = boto3.client(
            'ec2',
            aws_access_key_id=stsCreds['Credentials']['AccessKeyId'],
            aws_secret_access_key=stsCreds['Credentials']['SecretAccessKey'],
            aws_session_token=stsCreds['Credentials']['SessionToken'],
        )

    return client

# Visual modifications of xlsx to make it easier to read
def style(ws, value, row, column, color):
    ws.cell(row=row, column=column).value = value
    ws.cell(row=row, column=column).font = Font(bold=True,size=12)
    ws.cell(row=row, column=column).fill = PatternFill("solid", fgColor=color)
    ws.cell(row=row, column=column).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def sg_export(awsAccID, prefix, vpcId, fileName):
    client = login(awsAccID, prefix)
    securityGroups = client.describe_security_groups()

    try:
        wb = load_workbook(filename=fileName)
        ws = wb.create_sheet(title="Ec2 List")
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ec2 List"

    i = 0

    for sg in securityGroups['SecurityGroups']:
        if sg['VpcId'] == vpcId:
            style(ws, 'GroupName', i + 2, 2, "000000FF")
            ws.cell(row=2+i, column=3).value = sg['GroupName']
            style(ws, 'Description', 3 + i, 2, "000000FF")
            ws.cell(row=3+i, column=3).value = sg['Description']
            try:
                style(ws, 'Name', 4 + i, 2, "000000FF")
                for tag in sg['Tags']:
                    if tag['Key'] == 'Name':
                        ws.cell(row=4+i, column=3).value = tag['Value']
            except:
                style(ws, 'Name', 4 + i, 2, "000000FF")
                pass
            style(ws, 'Inbound Rules:', 5 + i, 2, "E59000")
            style(ws, 'Protocol', 6 + i, 2, "E59000")
            style(ws, 'Port Range', 6 + i, 3, "E59000")
            style(ws, 'Source', 6 + i, 4, "E59000")

            k = i
            for ruleIn in sg['IpPermissions']:
                ws.cell(row=7 + k, column=2).value = ruleIn['IpProtocol']
                if ruleIn['IpProtocol'] == "-1":
                    ws.cell(row=7 + k, column=3).value = ruleIn['IpProtocol']
                else:
                    if ruleIn['FromPort'] == ruleIn['ToPort']:
                        ws.cell(row=7 + k, column=3).value = ruleIn['FromPort']
                    else:
                        ws.cell(row=7 + k, column=3).value = str(ruleIn['FromPort']) + "-" + str(ruleIn['ToPort'])
                sourceList = ''
                for ip in ruleIn['IpRanges']:
                    sourceList = sourceList + ' ' + ip['CidrIp']
                    ws.cell(row=7 + k, column=4).value = sourceList
                for user in ruleIn['UserIdGroupPairs']:
                    sourceList = sourceList + ' ' + user['GroupId'] + '/' + user['UserId']
                    ws.cell(row=7 + k, column=4).value = sourceList
                k = k + 1

            style(ws, 'Outbound Rules:', 8 + k, 2, "E59000")
            style(ws, 'Protocol', 9 + k, 2, "E59000")
            style(ws, 'Port Range', 9 + k, 3, "E59000")
            style(ws, 'Source', 9 + k, 4, "E59000")
            for ruleOut in sg['IpPermissionsEgress']:
                ws.cell(row=10 + k, column=2).value = ruleOut['IpProtocol']
                if ruleOut['IpProtocol'] == "-1":
                    ws.cell(row=10 + k, column=3).value = ruleOut['IpProtocol']
                else:
                    if ruleOut['FromPort'] == ruleOut['ToPort']:
                        ws.cell(row=10 + k, column=3).value = ruleOut['FromPort']
                    else:
                        ws.cell(row=10 + k, column=3).value = str(ruleOut['FromPort']) + "-" + str(ruleOut['ToPort'])
                sourceList = ''
                for ip in ruleOut['IpRanges']:
                    sourceList = sourceList + ' ' + ip['CidrIp']
                    ws.cell(row=10 + k, column=4).value = sourceList
                for user in ruleOut['UserIdGroupPairs']:
                    sourceList = sourceList + ' ' + user['GroupId'] + '/' + user['UserId']
                    ws.cell(row=10 + k, column=4).value = sourceList
                k = k + 1

            i = k + 9

    wb.save(fileName)

# Main
sg_export(awsAccID, prefix, vpcId, fileName)